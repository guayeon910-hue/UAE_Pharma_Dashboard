"""DoH(아부다비 보건부) 및 DHA(두바이 보건국) 참조 가격 리스트 크롤러.

타겟:
  DoH: https://www.doh.gov.ae/en/resources/Circulars (참조 가격 Excel .ashx)
  DHA: https://www.dha.gov.ae/en/HealthRegulationSector/DrugControl (약가표 XLSX)

수집 항목:
  성분명(INN), 원산지, 제조사, 로컬 에이전트,
  약국공급가(AED), 대중판매가(AED), POM 여부

결과는 Supabase uae_price_reference 테이블에 저장됩니다.
"""
from __future__ import annotations

import asyncio
import io
import re
from typing import Any

import httpx

# DoH 참조 가격 리스트 URL (Shafafiya 포털의 직접 파일 다운로드 링크로 변경)
DOH_PRICE_LIST_URL = "https://www.doh.gov.ae/-/media/Feature/shafifya/Prices/Drugs.ashx"
DOH_CIRCULARS_URL = "https://www.doh.gov.ae/-/media/Feature/shafifya/dictionary/Reference-Price-List-All-Phases.ashx"

# DHA 약가표 URL
DHA_PRICE_LIST_URL = (
    "https://www.dha.gov.ae/en/HealthRegulationSector/DrugControl"
)

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; UPharma-MarketBot/1.0)",
    "Accept": "text/html,application/xhtml+xml,*/*",
}

# 인메모리 캐시 (세션당 1회 다운로드)
_doh_price_cache: list[dict[str, Any]] | None = None
_dha_price_cache: list[dict[str, Any]] | None = None


def _fetch_excel_data_sync(urls: list[str]) -> bytes | None:
    """Playwright를 이용해 페이지에서 Excel/ASHX 파일을 찾아 다운로드."""
    import time
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None
        
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="en-US",
            viewport={"width": 1920, "height": 1080}
        )
        page = context.new_page()
        
        target_href = None
        for url in urls:
            if not url:
                continue
                
            # URL이 직접 파일 링크(.ashx, .xlsx)인 경우 페이지 로드 대신 즉시 GET 요청 시도
            if any(ext in url.lower() for ext in [".ashx", ".xlsx", ".xls"]):
                try:
                    resp = context.request.get(url, timeout=30000)
                    if resp.status == 200:
                        ct = resp.headers.get("content-type", "").lower()
                        body = resp.body()
                        if len(body) > 10000 and ("excel" in ct or "spreadsheet" in ct or "octet-stream" in ct or "application/vnd" in ct):
                            browser.close()
                            return body
                except Exception:
                    pass
                continue
                
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                # Cloudflare/WAF 방어막 통과를 위해 충분히 대기
                page.wait_for_timeout(5000)
                
                links = page.query_selector_all("a[href]")
                for link in links:
                    href = link.get_attribute("href")
                    if href and any(ext in href.lower() for ext in [".xlsx", ".xls", ".ashx", "price", "pricelist"]):
                        if href.startswith("http"):
                            target_href = href
                        elif href.startswith("/"):
                            from urllib.parse import urlparse
                            parsed = urlparse(url)
                            target_href = f"{parsed.scheme}://{parsed.netloc}{href}"
                        else:
                            from urllib.parse import urljoin
                            target_href = urljoin(url, href)
                        
                        # URL 공백 등 인코딩 보정
                        from urllib.parse import urlparse, quote, urlunparse
                        parsed_href = urlparse(target_href)
                        safe_path = quote(parsed_href.path)
                        target_href = urlunparse((parsed_href.scheme, parsed_href.netloc, safe_path, parsed_href.params, parsed_href.query, parsed_href.fragment))
                        
                        # 파일이 맞는지 GET 요청으로 확인
                        try:
                            resp = context.request.get(target_href, timeout=15000)
                            if resp.status == 200:
                                ct = resp.headers.get("content-type", "").lower()
                                body = resp.body()
                                if len(body) > 10000 and ("excel" in ct or "spreadsheet" in ct or "octet-stream" in ct or "application/vnd" in ct):
                                    browser.close()
                                    return body
                        except Exception:
                            pass
                        target_href = None # 아니면 다음 링크 시도
            except Exception:
                pass

        browser.close()
        return None


def _parse_price_excel(data: bytes, source_label: str) -> list[dict[str, Any]]:
    """Excel 바이너리를 파싱하여 가격 행 리스트 반환."""
    rows: list[dict[str, Any]] = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        
        col_aliases = {
            "inn": ["inn", "active ingredient", "generic name", "innname", "active_ingredient", "generic"],
            "trade_name": ["trade name", "brand", "product name", "tradename", "package name"],
            "manufacturer": ["manufacturer", "mfr", "company"],
            "origin": ["origin", "country", "country of origin"],
            "agent": ["agent", "local agent", "distributor"],
            "pharmacy_price": ["pharmacy price", "pharmacy", "pharm price", "supply price"],
            "public_price": ["public price", "retail", "selling price", "mrp"],
            "pom": ["pom", "prescription", "rx"],
            "dosage_form": ["form", "dosage form", "formulation"],
            "strength": ["strength", "dose", "concentration"],
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_idx: dict[str, int] = {}
            
            for row in ws.iter_rows(values_only=True):
                # 헤더 찾기
                if not header_idx:
                    row_strs = [str(c or "").strip().lower() for c in row]
                    found_any = False
                    temp_idx = {}
                    for std_key, aliases in col_aliases.items():
                        for alias in aliases:
                            for hi, h in enumerate(row_strs):
                                if alias in h:
                                    temp_idx[std_key] = hi
                                    found_any = True
                                    break
                            if std_key in temp_idx:
                                break
                    
                    if "inn" in temp_idx or "trade_name" in temp_idx:
                        header_idx = temp_idx
                    continue

                # 데이터 파싱
                def _cell(key: str) -> str:
                    idx = header_idx.get(key)
                    if idx is not None and idx < len(row):
                        return str(row[idx] or "").strip()
                    return ""

                inn_val = _cell("inn")
                if not inn_val or inn_val.lower() in ("inn", "active ingredient", "generic name", "none", ""):
                    continue

                rows.append({
                    "inn":            inn_val,
                    "trade_name":     _cell("trade_name"),
                    "manufacturer":   _cell("manufacturer"),
                    "origin":         _cell("origin"),
                    "agent":          _cell("agent"),
                    "dosage_form":    _cell("dosage_form"),
                    "strength":       _cell("strength"),
                    "pharmacy_price_aed": _parse_aed(_cell("pharmacy_price")),
                    "public_price_aed":   _parse_aed(_cell("public_price")),
                    "pom":            "yes" in _cell("pom").lower() or _cell("pom") == "1",
                    "source":         source_label,
                })

    except Exception as e:
        import traceback
        traceback.print_exc()
        pass
    return rows


def _parse_aed(val: str) -> float | None:
    """'AED 12.50', '12.5', '12,500' 형식에서 float 추출."""
    if not val:
        return None
    cleaned = re.sub(r"[^\d.]", "", val.replace(",", ""))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


async def fetch_doh_prices() -> list[dict[str, Any]]:
    """DoH 참조 가격 리스트 다운로드 및 파싱."""
    global _doh_price_cache
    if _doh_price_cache is not None:
        return _doh_price_cache

    data = await asyncio.to_thread(_fetch_excel_data_sync, [DOH_PRICE_LIST_URL, DOH_CIRCULARS_URL])

    if data:
        _doh_price_cache = _parse_price_excel(data, "DoH Abu Dhabi")
        return _doh_price_cache

    # Jina AI 폴백 — 텍스트에서 가격 추출
    try:
        jina_url = f"https://r.jina.ai/{DOH_PRICE_LIST_URL}"
        async with httpx.AsyncClient(timeout=20.0) as client:
            resp = await client.get(jina_url)
            if resp.status_code == 200:
                _doh_price_cache = _parse_text_prices(resp.text, "DoH Abu Dhabi (Jina)")
                return _doh_price_cache
    except Exception:
        pass

    _doh_price_cache = []
    return _doh_price_cache


async def fetch_dha_prices() -> list[dict[str, Any]]:
    """DHA 두바이 약가표 다운로드 및 파싱."""
    global _dha_price_cache
    if _dha_price_cache is not None:
        return _dha_price_cache

    data = await asyncio.to_thread(_fetch_excel_data_sync, [DHA_PRICE_LIST_URL])
    
    if data:
        _dha_price_cache = _parse_price_excel(data, "DHA Dubai")
        return _dha_price_cache

    _dha_price_cache = []
    return _dha_price_cache


def _parse_text_prices(text: str, source: str) -> list[dict[str, Any]]:
    """텍스트(마크다운/HTML)에서 AED 가격 패턴 추출."""
    rows: list[dict[str, Any]] = []
    aed_pattern = re.compile(r"([A-Za-z\s\-]+)\s+(?:AED\s*)?([\d,]+\.?\d*)\s+(?:AED\s*)?([\d,]+\.?\d*)?")
    for line in text.split("\n"):
        if "error 404" in line.lower() or "not found" in line.lower():
            continue
        m = aed_pattern.search(line)
        if m:
            inn = m.group(1).strip()
            if inn.lower() in ("error", "page not found", "not found"):
                continue
            rows.append({
                "inn": inn,
                "pharmacy_price_aed": _parse_aed(m.group(2)),
                "public_price_aed": _parse_aed(m.group(3)) if m.group(3) else None,
                "source": source,
            })
    return rows[:100]


async def get_price_context_for_inn(inn: str) -> str | None:
    """INN에 대한 DoH/DHA 가격 컨텍스트 문자열 생성 (Claude 프롬프트용)."""
    if not inn or not inn.strip():
        return None

    inn_parts = [p.strip().lower() for p in re.split(r"[/+,]", inn) if p.strip()]
    
    from utils.db import get_client
    sb = get_client()
    try:
        # INN 이름이나 제품명에 첫 번째 단어가 포함된 모든 가격 데이터 검색
        first_part = inn_parts[0] if inn_parts else inn
        res = sb.table("uae_price_reference").select("*").or_(f"inn_name.ilike.%{first_part}%,trade_name.ilike.%{first_part}%").execute()
        all_rows = res.data or []
    except Exception as e:
        import logging
        logging.getLogger("uae.crawler").error(f"Failed to fetch price context from Supabase: {e}")
        return None

    if not all_rows:
        return None

    matched: list[dict[str, Any]] = []
    for row in all_rows:
        row_inn = (row.get("inn_name") or "").lower()
        row_trade = (row.get("trade_name") or "").lower()
        if any(part in row_inn for part in inn_parts) or any(part in row_trade for part in inn_parts):
            matched.append(row)

    if not matched:
        return None

    lines = [f"참조 가격 데이터 ({len(matched)}건 매칭, 성분: {inn}):"]
    for r in matched[:8]:
        ph = r.get("pharmacy_price_aed")
        pub = r.get("public_price_aed")
        ph_str = f"약국공급가 AED {ph:.2f}" if ph else ""
        pub_str = f"대중판매가 AED {pub:.2f}" if pub else ""
        price_str = " / ".join(filter(None, [ph_str, pub_str])) or "가격 미확인"
        pom = "처방전 의약품(POM)" if r.get("is_pom") else "일반의약품(OTC)"
        origin = r.get("origin_country", "")
        mfr = r.get("manufacturer", "")
        trade = r.get("trade_name", "")
        parts = [f"- {trade or r.get('inn_name', '')}"]
        if mfr:
            parts.append(f"제조사: {mfr}")
        if origin:
            parts.append(f"원산지: {origin}")
        parts.append(price_str)
        parts.append(pom)
        parts.append(f"[출처: {r.get('source_label', '')}]")
        lines.append(" | ".join(parts))

    return "\n".join(lines)


async def save_prices_to_supabase(rows: list[dict[str, Any]]) -> int:
    """수집된 가격 데이터를 Supabase uae_price_reference 테이블에 저장."""
    if not rows:
        return 0
    try:
        from utils.db import get_client
        from datetime import datetime, timezone
        sb = get_client()
        seen = set()
        to_upsert = []
        for r in rows:
            inn = r.get("inn", "")
            trade = r.get("trade_name", "")
            source = r.get("source", "")
            key = (inn, trade, source)
            
            # DB의 unique constraint에 맞춰 메모리상 중복 제거
            if key in seen:
                continue
            seen.add(key)
            
            to_upsert.append({
                "inn_name":            inn,
                "trade_name":          trade,
                "manufacturer":        r.get("manufacturer", ""),
                "origin_country":      r.get("origin", ""),
                "local_agent":         r.get("agent", ""),
                "dosage_form":         r.get("dosage_form", ""),
                "strength":            r.get("strength", ""),
                "pharmacy_price_aed":  r.get("pharmacy_price_aed"),
                "public_price_aed":    r.get("public_price_aed"),
                "is_pom":              r.get("pom", True),
                "source_label":        source,
                "crawled_at":          datetime.now(timezone.utc).isoformat(),
            })
            
        total_upserted = 0
        # 최대 1000건씩 쪼개어 upsert
        for i in range(0, len(to_upsert), 1000):
            chunk = to_upsert[i:i+1000]
            try:
                result = sb.table("uae_price_reference").upsert(
                    chunk, on_conflict="inn_name,trade_name,source_label"
                ).execute()
                total_upserted += len(result.data or [])
            except Exception as e:
                import traceback
                traceback.print_exc()
        return total_upserted
    except Exception as e:
        import traceback
        traceback.print_exc()
        return 0
